import csv

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Avg, Max, Count, OuterRef, Subquery
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST

from accounts.utils import admin_required, teacher_required
from courses.models import Course, Enrollment
from content.models import Lesson, TopicTemplate, LessonProgress
from quiz.models import Exam, Attempt, Answer
from payments.models import PurchaseRequest
from messaging.models import Conversation

from .forms import (
    CourseForm,
    LessonForm,
    LessonFileForm,
    LessonVideoForm,
    ExamForm,
    QuestionForm,
    LiveLessonForm,
    EnrollStudentForm,
    AddTopicLessonsForm,
)

User = get_user_model()


# =========================
# TEACHER HOME
# =========================
@login_required
@teacher_required
def teacher_home(request):
    if request.user.role == "ADMIN":
        courses = Course.objects.all()
    else:
        courses = Course.objects.filter(owner=request.user)
    return render(request, "teacher/home.html", {"courses": courses})


# =========================
# COURSE (ADMIN)
# =========================
@login_required
@admin_required
def course_new(request):
    if request.method == "POST":
        form = CourseForm(request.POST)
        if form.is_valid():
            course = form.save(commit=False)
            course.owner = request.user
            course.save()
            return redirect("teacher_home")
    else:
        form = CourseForm()

    return render(request, "teacher/course_new.html", {"form": form})


# =========================
# LESSON (MANUAL)
# =========================
@login_required
@teacher_required
def lesson_new(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST":
        form = LessonForm(request.POST)
        if form.is_valid():
            lesson = form.save(commit=False)
            lesson.course = course
            lesson.save()
            return redirect("course_detail", course_id=course.id)
    else:
        form = LessonForm()

    return render(request, "teacher/lesson_new.html", {"form": form, "course": course})


@login_required
@teacher_required
def edit_lesson(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)
    course = lesson.course

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST":
        form = LessonForm(request.POST, instance=lesson)
        if form.is_valid():
            form.save()
            messages.success(request, "Ders güncellendi.")
            return redirect("course_detail", course_id=course.id)
    else:
        form = LessonForm(instance=lesson)

    return render(request, "teacher/lesson_edit.html", {"form": form, "lesson": lesson, "course": course})


@login_required
@teacher_required
@require_POST
def delete_lesson(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)
    course = lesson.course

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    lesson.delete()
    messages.success(request, "Ders silindi.")
    return redirect("course_detail", course_id=course.id)


@login_required
@teacher_required
@require_POST
def reorder_lessons(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    lessons = Lesson.objects.filter(course=course)
    updated = 0

    for l in lessons:
        key = f"order_{l.id}"
        val = request.POST.get(key)
        if val and val.isdigit():
            new_order = int(val)
            if l.order != new_order:
                l.order = new_order
                l.save(update_fields=["order"])
                updated += 1

    messages.success(request, f"Sıralama güncellendi. ({updated} ders)")
    return redirect("course_detail", course_id=course.id)


# =========================
# FILE / VIDEO UPLOAD
# =========================
@login_required
@teacher_required
def file_new(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)

    if request.user.role != "ADMIN" and lesson.course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST":
        form = LessonFileForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.save(commit=False)
            f.lesson = lesson
            f.save()
            return redirect("course_detail", course_id=lesson.course.id)
    else:
        form = LessonFileForm()

    return render(request, "teacher/file_new.html", {"form": form, "lesson": lesson})


@login_required
@teacher_required
def video_new(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)

    if request.user.role != "ADMIN" and lesson.course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST":
        form = LessonVideoForm(request.POST, request.FILES)
        if form.is_valid():
            v = form.save(commit=False)
            v.lesson = lesson
            v.save()
            return redirect("course_detail", course_id=lesson.course.id)
    else:
        form = LessonVideoForm()

    return render(request, "teacher/video_new.html", {"form": form, "lesson": lesson})


# =========================
# EXAM
# =========================
@login_required
@teacher_required
def exam_new(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST":
        form = ExamForm(request.POST)
        if form.is_valid():
            exam = form.save(commit=False)
            exam.course = course
            exam.save()
            return redirect("course_detail", course_id=course.id)
    else:
        form = ExamForm()

    return render(request, "teacher/exam_new.html", {"form": form, "course": course})


@login_required
@teacher_required
def question_new(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)

    if request.user.role != "ADMIN" and exam.course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST":
        form = QuestionForm(request.POST)
        if form.is_valid():
            qobj = form.save(commit=False)
            qobj.exam = exam
            qobj.save()
            return redirect("teacher_question_new", exam_id=exam.id)
    else:
        form = QuestionForm()

    questions = exam.questions.all()
    return render(request, "teacher/question_new.html", {"form": form, "exam": exam, "questions": questions})


# =========================
# LIVE LESSON
# =========================
@login_required
@teacher_required
def live_lesson_new(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST":
        form = LiveLessonForm(request.POST)
        if form.is_valid():
            ll = form.save(commit=False)
            ll.course = course
            ll.save()
            return redirect("course_detail", course_id=course.id)
    else:
        form = LiveLessonForm()

    return render(request, "teacher/live_lesson_new.html", {"form": form, "course": course})


# =========================
# COURSE STUDENTS (Teacher/Admin view)
# - GET: Teacher/Admin görebilir
# - POST (add/remove): sadece Admin
# =========================
@login_required
@teacher_required
def course_students(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    # Teacher sadece kendi kursu, admin her kurs
    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    can_manage = request.user.role == "ADMIN"
    form = EnrollStudentForm(request.POST or None) if can_manage else None

    # Admin değilse POST yasak
    if request.method == "POST" and not can_manage:
        return render(request, "courses/forbidden.html", status=403)

    if request.method == "POST" and "add_student" in request.POST:
        if form and form.is_valid():
            uname = form.cleaned_data["username"]
            student = User.objects.get(username=uname)
            Enrollment.objects.get_or_create(course=course, user=student)
            messages.success(request, f"{student.username} kursa eklendi.")
            return redirect("teacher_course_students", course_id=course.id)

    if request.method == "POST" and "remove_student" in request.POST:
        enrollment_id = request.POST.get("enrollment_id")
        Enrollment.objects.filter(id=enrollment_id, course=course).delete()
        messages.success(request, "Öğrenci kurstan çıkarıldı.")
        return redirect("teacher_course_students", course_id=course.id)

    enrollments = Enrollment.objects.filter(course=course).select_related("user").order_by("user__username")
    student_ids = [e.user_id for e in enrollments]

    total_lessons = course.lessons.count()

    # Ders ilerleme (tamamlanan sayısı + son tamamlama)
    lp = (
        LessonProgress.objects
        .filter(lesson__course=course, user_id__in=student_ids, completed=True)
        .values("user_id")
        .annotate(done=Count("id"), last_done=Max("completed_at"))
    )
    lesson_map = {x["user_id"]: x for x in lp}

    # Sınav özet (tamamlanan attempt sayısı + ortalama puan + son sınav)
    ap = (
        Attempt.objects
        .filter(exam__course=course, user_id__in=student_ids, finished_at__isnull=False)
        .values("user_id")
        .annotate(exam_done=Count("id"), avg_score=Avg("score"), last_exam=Max("finished_at"))
    )
    exam_map = {x["user_id"]: x for x in ap}

    rows = []
    for e in enrollments:
        u = e.user
        done = lesson_map.get(u.id, {}).get("done", 0)
        last_done = lesson_map.get(u.id, {}).get("last_done")
        percent = int((done / total_lessons) * 100) if total_lessons else 0

        exam_done = exam_map.get(u.id, {}).get("exam_done", 0)
        avg_score = exam_map.get(u.id, {}).get("avg_score") or 0
        last_exam = exam_map.get(u.id, {}).get("last_exam")

        rows.append({
            "enrollment": e,
            "user": u,
            "done": done,
            "total_lessons": total_lessons,
            "percent": percent,
            "last_done": last_done,
            "exam_done": exam_done,
            "avg_score": round(float(avg_score), 2),
            "last_exam": last_exam,
        })

    return render(request, "teacher/course_students.html", {
        "course": course,
        "form": form,
        "can_manage": can_manage,
        "rows": rows,
        "student_count": len(rows),
    })


# =========================
# READY LESSONS (ADMIN)
# =========================
@login_required
@admin_required
def add_ready_lessons(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    form = AddTopicLessonsForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        grade = form.cleaned_data["grade"]
        subject = form.cleaned_data["subject"]

        parents = TopicTemplate.objects.filter(
            grade=grade, subject=subject, parent__isnull=True
        ).order_by("order", "id")

        created = 0

        for p in parents:
            _, created_parent = Lesson.objects.get_or_create(
                course=course,
                title=p.title,
                defaults={"order": p.order * 100},
            )
            if created_parent:
                created += 1

            children = p.children.all().order_by("order", "id")
            for c in children:
                child_title = f"— {c.title}"
                _, created_child = Lesson.objects.get_or_create(
                    course=course,
                    title=child_title,
                    defaults={"order": (p.order * 100) + c.order},
                )
                if created_child:
                    created += 1

        try:
            course.grade = grade.number
            course.save()
        except Exception:
            pass

        messages.success(request, f"{grade.number}. sınıf {subject.name} için {created} ders eklendi.")
        return redirect("course_detail", course_id=course.id)

    return render(request, "teacher/add_ready_lessons.html", {"course": course, "form": form})


# =========================
# PAYMENTS (ADMIN)
# =========================
@login_required
@admin_required
def purchase_requests(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    items = PurchaseRequest.objects.filter(course=course).select_related("user").order_by("-created_at")
    return render(request, "teacher/purchase_requests.html", {"course": course, "items": items})


@login_required
@admin_required
def purchase_request_approve(request, pr_id):
    pr = get_object_or_404(PurchaseRequest, id=pr_id)
    course = pr.course

    pr.status = PurchaseRequest.Status.APPROVED
    pr.save()

    Enrollment.objects.get_or_create(course=course, user=pr.user)
    messages.success(request, f"{pr.user.username} için kurs erişimi açıldı.")
    return redirect(f"/teacher/course/{course.id}/payments/")


# =========================
# EXAM RESULTS
# =========================
@login_required
@teacher_required
def exam_results(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)

    if request.user.role != "ADMIN" and exam.course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    q = (request.GET.get("q") or "").strip()
    finished_only = request.GET.get("finished") == "1"
    export = request.GET.get("export")  # csv / xlsx

    attempts = Attempt.objects.filter(exam=exam).select_related("user").order_by("-started_at")

    if finished_only:
        attempts = attempts.filter(finished_at__isnull=False)

    if q:
        attempts = attempts.filter(
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )

    if export == "csv":
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="exam_{exam.id}_results.csv"'
        resp.write("\ufeff")
        w = csv.writer(resp)
        w.writerow(["Ogrenci", "Baslangic", "Bitis", "Puan", "Dogru", "Yanlis"])
        for a in attempts:
            w.writerow([a.user.username, a.started_at, a.finished_at or "", a.score if a.finished_at else "", a.correct_count, a.wrong_count])
        return resp

    if export == "xlsx":
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Sonuclar"
        ws.append(["Ogrenci", "Baslangic", "Bitis", "Puan", "Dogru", "Yanlis"])

        for a in attempts:
            ws.append([a.user.username, str(a.started_at), str(a.finished_at) if a.finished_at else "", a.score if a.finished_at else "", a.correct_count, a.wrong_count])

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22

        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="exam_{exam.id}_results.xlsx"'
        wb.save(resp)
        return resp

    finished_qs = attempts.filter(finished_at__isnull=False)
    finished_count = finished_qs.count()
    total_count = attempts.count()
    avg_score = round(sum(a.score for a in finished_qs) / finished_count, 2) if finished_count else 0

    finished_attempts = attempts.filter(finished_at__isnull=False)

    last_score_subq = (
        Attempt.objects
        .filter(exam=exam, user_id=OuterRef("user_id"), finished_at__isnull=False)
        .order_by("-finished_at")
        .values("score")[:1]
    )

    student_summary = (
        finished_attempts
        .values("user_id", "user__username")
        .annotate(
            attempt_count=Count("id"),
            avg_score=Avg("score"),
            best_score=Max("score"),
            last_finished=Max("finished_at"),
            last_score=Subquery(last_score_subq),
        )
        .order_by("-avg_score", "user__username")
    )

    student_count = student_summary.count()

    return render(request, "teacher/exam_results.html", {
        "exam": exam,
        "attempts": attempts,
        "avg_score": avg_score,
        "finished_count": finished_count,
        "total_count": total_count,
        "q": q,
        "finished_only": finished_only,
        "student_summary": student_summary,
        "student_count": student_count,
    })


@login_required
@teacher_required
def attempt_detail(request, attempt_id):
    attempt = get_object_or_404(Attempt, id=attempt_id)
    exam = attempt.exam

    if request.user.role != "ADMIN" and exam.course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    answers = Answer.objects.filter(attempt=attempt).select_related("question").order_by("question__order", "question__id")
    return render(request, "teacher/attempt_detail.html", {"attempt": attempt, "exam": exam, "answers": answers})


# =========================
# COURSE PROGRESS (Teacher/Admin) + q + csv + avg_percent + low
# =========================
@login_required
@teacher_required
def course_progress(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    total_lessons = course.lessons.count()

    q = (request.GET.get("q") or "").strip()
    only_low = request.GET.get("low") == "1"
    export = request.GET.get("export")  # csv

    enrollments = Enrollment.objects.filter(course=course).select_related("user").order_by("user__username")

    if q:
        enrollments = enrollments.filter(
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )

    student_ids = [e.user_id for e in enrollments]

    prog = (
        LessonProgress.objects
        .filter(lesson__course=course, user_id__in=student_ids, completed=True)
        .values("user_id")
        .annotate(done=Count("id"), last_done=Max("completed_at"))
    )
    prog_map = {p["user_id"]: p for p in prog}

    rows = []
    for e in enrollments:
        u = e.user
        done = prog_map.get(u.id, {}).get("done", 0)
        last_done = prog_map.get(u.id, {}).get("last_done")
        percent = int((done / total_lessons) * 100) if total_lessons else 0
        rows.append({"user": u, "done": done, "total": total_lessons, "percent": percent, "last_done": last_done})

    if only_low:
        rows = [r for r in rows if r["percent"] < 50]

    avg_percent = round(sum(r["percent"] for r in rows) / len(rows), 2) if rows else 0

    if export == "csv":
        resp = HttpResponse(content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="course_{course.id}_progress.csv"'
        resp.write("\ufeff")
        w = csv.writer(resp)
        w.writerow(["Ogrenci", "Tamamlanan", "Toplam", "Yuzde", "Son Tamamlama"])
        for r in rows:
            w.writerow([r["user"].username, r["done"], r["total"], r["percent"], r["last_done"] or ""])
        return resp

    return render(request, "teacher/course_progress.html", {
        "course": course,
        "rows": rows,
        "total_lessons": total_lessons,
        "only_low": only_low,
        "q": q,
        "avg_percent": avg_percent,
    })


# =========================
# COURSE PROGRESS STUDENT DETAIL
# =========================
@login_required
@teacher_required
def course_progress_student_detail(request, course_id, student_id):
    course = get_object_or_404(Course, id=course_id)

    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    enrollment = Enrollment.objects.filter(course=course, user_id=student_id).select_related("user").first()
    if not enrollment:
        return render(request, "courses/forbidden.html", status=403)

    student = enrollment.user

    lessons = course.lessons.order_by("order", "id")
    total = lessons.count()

    prog_qs = LessonProgress.objects.filter(user=student, lesson__course=course)
    prog_map = {p.lesson_id: p for p in prog_qs}
    completed_ids = {p.lesson_id for p in prog_qs if p.completed}

    done = len(completed_ids)
    percent = int((done / total) * 100) if total else 0

    rows = []
    for l in lessons:
        p = prog_map.get(l.id)
        rows.append({
            "lesson": l,
            "completed": bool(p and p.completed),
            "completed_at": p.completed_at if (p and p.completed) else None,
        })

    return render(request, "teacher/course_progress_student_detail.html", {
        "course": course,
        "student": student,
        "rows": rows,
        "done": done,
        "total": total,
        "percent": percent,
    })
    
@login_required
@teacher_required
def teacher_message_student(request, course_id, student_id):
    course = get_object_or_404(Course, id=course_id)

    # Yetki: teacher kendi kursu / admin her kurs
    if request.user.role != "ADMIN" and course.owner_id != request.user.id:
        return render(request, "courses/forbidden.html", status=403)

    # Öğrenci kursa kayıtlı mı?
    if not Enrollment.objects.filter(course=course, user_id=student_id).exists():
        return render(request, "courses/forbidden.html", status=403)

    student = get_object_or_404(User, id=student_id)

    # Konuşmayı course owner öğretmenine bağlayalım (tek thread)
    teacher_user = course.owner

    convo, _ = Conversation.objects.get_or_create(
        course=course,
        student=student,
        teacher=teacher_user
    )

    return redirect("teacher_chat", conversation_id=convo.id)
